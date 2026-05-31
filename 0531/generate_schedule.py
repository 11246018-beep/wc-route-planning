import pandas as pd
import numpy as np
from math import radians, cos, sin, asin, sqrt
import requests
import time

# Configuration
INPUT_FILE = 'maintenance_data_v2.xlsx'
OUTPUT_FILE = 'Weekly_Schedule.xlsx'

# OSRM Configuration
OSRM_BASE_URL = 'http://router.project-osrm.org/route/v1/driving'
OSRM_CACHE = {} # Simple in-memory cache
MIN_OSRM_DIST_KM = 5.0 # Optimization: Skip OSRM for < 5km

# Staff Configuration
STAFF_CONFIG = [
    # PZ Staff cascading south
    {'id': 'S01', 'team': 'PZ', 'zones': ['PZ-A', 'PZ-B', 'WG-D']}, # Added WG-D (Guanyin/Taoyuan overlap)
    {'id': 'S02', 'team': 'PZ', 'zones': ['PZ-A', 'PZ-B', 'WG-D']},
    {'id': 'S03', 'team': 'PZ', 'zones': ['PZ-A', 'PZ-B', 'WG-D']},
    {'id': 'S04', 'team': 'PZ', 'zones': ['PZ-A', 'PZ-B', 'WG-D']},
    {'id': 'S05', 'team': 'PZ', 'zones': ['PZ-B', 'PZ-C', 'PZ-A']}, # Added PZ-A to support S01-04
    {'id': 'S06', 'team': 'PZ', 'zones': ['PZ-B', 'PZ-C', 'PZ-A']},
    {'id': 'S07', 'team': 'PZ', 'zones': ['PZ-B', 'PZ-C', 'PZ-A']},
    {'id': 'S08', 'team': 'PZ', 'zones': ['PZ-D', 'PZ-C']}, # Added PZ-C (Miaoli helping Hsinchu)
    {'id': 'S09', 'team': 'PZ', 'zones': ['PZ-D', 'PZ-C']},
    {'id': 'S10', 'team': 'PZ', 'zones': ['PZ-D', 'PZ-C']},
    {'id': 'S11', 'team': 'PZ', 'zones': ['PZ-C', 'PZ-D']}, # C supports D
    {'id': 'S12', 'team': 'PZ', 'zones': ['PZ-C', 'PZ-D']},
    # Wugu Staff supporting North - Optimized for WG-C coverage
    # Removed PZ-A and WG-D to force them to focus on WG-C/WG-B (Remote) -> Added PZ-A back as secondary
    {'id': 'S13', 'team': 'WG', 'zones': ['WG-C', 'WG-B', 'WG-A', 'PZ-A']}, 
    {'id': 'S14', 'team': 'WG', 'zones': ['WG-C', 'WG-B', 'WG-A', 'PZ-A']}
]

DAILY_LIMIT_MINS = 540 # Strict limit requested by user
AVG_SPEED_KMPH = 25  # Conservative city speed (Fallback)

def haversine(lon1, lat1, lon2, lat2):
    """
    Calculate the great circle distance in kilometers between two points 
    on the earth (specified in decimal degrees)
    """
    if pd.isna(lon1) or pd.isna(lat1) or pd.isna(lon2) or pd.isna(lat2):
        return 0
    # convert decimal degrees to radians 
    lon1, lat1, lon2, lat2 = map(radians, [lon1, lat1, lon2, lat2])
    # haversine formula 
    dlon = lon2 - lon1 
    dlat = lat2 - lat1 
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * asin(sqrt(a)) 
    r = 6371 # Radius of earth in kilometers
    return c * r

def estimate_travel_time_fallback(km):
    """Fallback if OSRM fails or distance is short"""
    if km < 0.1: return 2 
    speed = AVG_SPEED_KMPH
    if km > 10:
        speed = 60
    hours = km / speed
    mins = hours * 60
    return max(mins, 3) 

def get_osrm_duration(lon1, lat1, lon2, lat2):
    """
    Get duration in minutes from OSRM API.
    Returns: duration_mins (float) or None if failed.
    """
    # 1. Round coordinates to increase cache hits (4 decimal places ~ 11m)
    r_lon1, r_lat1 = round(lon1, 4), round(lat1, 4)
    r_lon2, r_lat2 = round(lon2, 4), round(lat2, 4)
    
    # 2. Check Cache
    cache_key = (r_lon1, r_lat1, r_lon2, r_lat2)
    if cache_key in OSRM_CACHE:
        return OSRM_CACHE[cache_key]
        
    # 3. Check Short Distance Optimization
    # Calculate Haversine first
    dist_km = haversine(lon1, lat1, lon2, lat2)
    if dist_km < MIN_OSRM_DIST_KM:
        val = estimate_travel_time_fallback(dist_km)
        OSRM_CACHE[cache_key] = val
        return val

    # 4. Call API
    try:
        url = f"{OSRM_BASE_URL}/{lon1},{lat1};{lon2},{lat2}?overview=false"
        print(".", end="", flush=True) # Progress Indicator
        response = requests.get(url, timeout=0.5) # Aggressive timeout
        if response.status_code == 200:
            data = response.json()
            if 'routes' in data and len(data['routes']) > 0:
                duration_seconds = data['routes'][0]['duration']
                val = duration_seconds / 60.0 # Convert to minutes
                OSRM_CACHE[cache_key] = val
                return val
    except Exception as e:
        pass
        
    # 5. Fallback
    val = estimate_travel_time_fallback(dist_km)
    OSRM_CACHE[cache_key] = val
    return val

def get_county(address):
    """Extract first 3 characters as county/city. Treat Hsinchu County/City as one."""
    if not isinstance(address, str) or len(address) < 3:
        return 'Unknown'
    county = address[:3]
    if '新竹' in county:
        return '新竹' # Merged
    return county

def get_zone(sheet, addr):
    addr = str(addr)
    # Pingzhen Logic
    if '平鎮' in sheet:
        # Priority 1: Explicit Zone Labels
        a_sub = ['大園', '蘆竹', '觀音', '桃園區', '龜山']
        b_sub = ['中壢', '平鎮', '八德', '龍潭', '楊梅', '新屋']
        c_sub = ['新竹', '竹北', '湖口', '新豐', '寶山', '竹東', '芎林', '新埔', '關西', '峨眉']
        d_sub = ['苗栗', '新北', '臺北', '台北', '頭份', '竹南', '公館', '造橋']

        if any(k in addr for k in c_sub): return 'PZ-C'
        if any(k in addr for k in d_sub): return 'PZ-D'
        if any(k in addr for k in b_sub): return 'PZ-B'
        if any(k in addr for k in a_sub): return 'PZ-A'
        
        if '桃園' in addr and '區' not in addr: return 'PZ-A' # Fallback
        
        return 'PZ-B' # Ultimate Default (Central Area)
    # Wugu Logic
    elif '五股' in sheet:
        if any(k in addr for k in ['北投', '淡水', '八里', '三芝', '金山', '石門']): return 'WG-B'
        if any(k in addr for k in ['基隆', '汐止', '瑞芳', '七堵', '安樂', '中山', '仁愛', '信義', '中正']): return 'WG-C'
        if any(k in addr for k in ['桃園', '龜山']): return 'WG-D'
        return 'WG-A' # Default
    return 'Unknown'

def load_tasks():
    print("Loading tasks...")
    all_tasks = []
    
    # Process each sheet
    for sheet in ['平鎮', '平鎮週清2', '五股', '五股週清2']:
        try:
            df = pd.read_excel(INPUT_FILE, sheet_name=sheet)
            
            # Identify columns
            id_col = 'ID'
            addr_col = next((c for c in df.columns if '地址' in str(c) or '地點' in str(c)), None)
            lat_col = next((c for c in df.columns if '緯度' in str(c)), None)
            lon_col = next((c for c in df.columns if '經度' in str(c)), None)
            work_col = next((c for c in df.columns if '維護時間' in str(c)), None)
            
            df = df.dropna(subset=[id_col])
            
            if lat_col and lon_col:
                df = df.dropna(subset=[lat_col, lon_col])
            
            for _, row in df.iterrows():
                # Handle comma separated IDs
                ids = str(row[id_col]).split(',')
                
                total_work = row[work_col] if work_col else 30 # Default 30 if missing
                
                # Determine Zone
                if '工作區' in df.columns and pd.notna(row['工作區']):
                    z_raw = str(row['工作區'])
                    if 'PZ-A' in z_raw: zone = 'PZ-A'
                    elif 'PZ-B' in z_raw: zone = 'PZ-B'
                    elif 'PZ-C' in z_raw: zone = 'PZ-C'
                    elif 'PZ-D' in z_raw: zone = 'PZ-D'
                    else: zone = get_zone(sheet, row[addr_col])
                else:
                    zone = get_zone(sheet, row[addr_col])
                
                all_tasks.append({
                    'ID_List': row[id_col],
                    'Address': row[addr_col],
                    'Lat': row[lat_col],
                    'Lon': row[lon_col],
                    'Work_Mins': total_work,
                    'Zone': zone,
                    'Sheet': sheet,
                    'Assigned': False
                })
                
        except Exception as e:
            print(f"Error loading {sheet}: {e}")
            
    print(f"Loaded {len(all_tasks)} stops.")
    return all_tasks

def generate_schedule():
    tasks = load_tasks()
    schedule = []
    
    # Bases
    base_coords = {'PZ': (121.22683, 24.90679), 'WG': (121.44141, 25.07055)} # Default fallback
    base_info = {
        'PZ': {'lon': 121.22683, 'lat': 24.90679, 'county': '桃園市'}, 
        'WG': {'lon': 121.44141, 'lat': 25.07055, 'county': '新北市'}
    }

    try:
        df_wh = pd.read_excel(INPUT_FILE, sheet_name='倉位地點')
        name_col = df_wh.columns[0]
        lon_col = next((c for c in df_wh.columns if '經度' in str(c)), None)
        lat_col = next((c for c in df_wh.columns if '緯度' in str(c)), None)
        
        if lon_col and lat_col:
            for _, row in df_wh.iterrows():
                name = str(row[name_col])
                lat = row[lat_col]
                lon = row[lon_col]
                if '平鎮' in name:
                    base_coords['PZ'] = (lon, lat)
                    base_info['PZ'] = {'lon': lon, 'lat': lat, 'county': '桃園市'} 
                elif '五股' in name:
                    base_coords['WG'] = (lon, lat)
                    base_info['WG'] = {'lon': lon, 'lat': lat, 'county': '新北市'}
            print(f"Loaded Base Coords: {base_coords}")
    except Exception as e:
        print(f"Warning: Could not load warehouse sheet, using defaults. {e}")
    
    # Assign Pools to Staff
    # We round-robin assign tasks from Zone Pools to eligible Staff
    # Better approach: Iterate Days 1-6. Staff picks nearest task.
    
    full_schedule = []
    
    for day in range(1, 7):
        print(f"--- Scheduling Day {day} ---")
        
        # Reset Staff State for new Day
        staff_state = {}
        for s in STAFF_CONFIG:
            team = s['team']
            start_lon, start_lat = base_coords[team]
            start_county = base_info[team]['county']
            staff_state[s['id']] = {
                'curr_lon': start_lon,
                'curr_lat': start_lat,
                'curr_county': start_county,
                'time_used': 0,
                'tasks_done': 0
            }
            
        # --- Phase 4: Priority Assignment for Load Balancing ---
        
        # 4a. Priority for WG-C (Keelung) for S13, S14
        wg_priority_ids = ['S13', 'S14']
        wg_priority_zone = 'WG-C'
        
        p_change = True
        while p_change:
            p_change = False
            for s_id in wg_priority_ids:
                state = staff_state[s_id]
                if state['time_used'] >= DAILY_LIMIT_MINS: continue
                
                # Find eligible WG-C tasks
                candidates = []
                for i, t in enumerate(tasks):
                    if not t['Assigned'] and t['Zone'] == wg_priority_zone:
                            dist = haversine(state['curr_lon'], state['curr_lat'], t['Lon'], t['Lat'])
                            candidates.append((dist, i))
                
                candidates.sort(key=lambda x: x[0])
                
                # Assign nearest
                for _, idx in candidates[:1]: # Take top 1
                    t = tasks[idx]
                     # Routing
                    start_lon, start_lat = state['curr_lon'], state['curr_lat']
                    end_lon, end_lat = t['Lon'], t['Lat']
                    
                    leg = get_osrm_duration(start_lon, start_lat, end_lon, end_lat)
                    if leg is None: 
                        dist = haversine(start_lon, start_lat, end_lon, end_lat)
                        leg = estimate_travel_time_fallback(dist)
                    
                    work_time = t['Work_Mins']
                    total = state['time_used'] + leg + work_time
                    
                    # Return to base check
                    base_key = 'WG' # S13/14 are WG
                    base_lon, base_lat = base_coords[base_key]
                    ret = get_osrm_duration(end_lon, end_lat, base_lon, base_lat)
                    if ret is None: ret = estimate_travel_time_fallback(haversine(end_lon, end_lat, base_lon, base_lat))
                    
                    if total + ret <= DAILY_LIMIT_MINS:
                        tasks[idx]['Assigned'] = True
                        state['time_used'] = total
                        state['curr_lon'] = end_lon
                        state['curr_lat'] = end_lat
                        state['curr_county'] = get_county(t['Address']) # Or '基隆市'
                        state['tasks_done'] += 1
                        
                        full_schedule.append({
                            'Day': day,
                            'Staff_ID': s_id,
                            'Zone': t['Zone'],
                            'Seq': state['tasks_done'],
                            'Task_IDs': t['ID_List'],
                            'Address': t['Address'],
                            'Lat': t['Lat'],
                            'Lon': t['Lon'],
                            'Work_Mins': work_time,
                            'Travel_Mins': round(leg, 1),
                            'Total_Time': round(state['time_used'], 1),
                            'Job_Type': 'Primary (Priority)'
                        })
                        p_change = True
                        break
        
        # 4b. Identify Priority Group PZ-B for S05-S07 (REVERTED from PZ-C)
        priority_group_ids = ['S05', 'S06', 'S07']
        priority_zone = 'PZ-B' # Changing focus to Hsinchu (PZ-C)
        
        # 0. Run Priority Loop First (Greedy for PZ-B)
        # This loop runs until S05-S07 are full or no PZ-B tasks remain
        p_change = True
        while p_change:
            p_change = False
            for s_id in priority_group_ids:
                state = staff_state[s_id]
                if state['time_used'] >= DAILY_LIMIT_MINS: continue
                
                # Find eligible PZ-B tasks
                candidates = []
                for i, t in enumerate(tasks):
                    if not t['Assigned'] and t['Zone'] == priority_zone:
                            dist = haversine(state['curr_lon'], state['curr_lat'], t['Lon'], t['Lat'])
                            candidates.append((dist, i))
                
                candidates.sort(key=lambda x: x[0])
                
                # Assign nearest
                for _, idx in candidates[:1]: # Take top 1
                    t = tasks[idx]
                    
                    # Routing (Simplified copy of below)
                    start_lon, start_lat = state['curr_lon'], state['curr_lat']
                    end_lon, end_lat = t['Lon'], t['Lat']
                    
                    # Check Cross County
                    leg = get_osrm_duration(start_lon, start_lat, end_lon, end_lat)
                    if leg is None: 
                        dist = haversine(start_lon, start_lat, end_lon, end_lat)
                        leg = estimate_travel_time_fallback(dist)
                    
                    work_time = t['Work_Mins']
                    total = state['time_used'] + leg + work_time
                    
                    # Return to base check
                    base_key = 'PZ'
                    base_lon, base_lat = base_coords[base_key]
                    ret = get_osrm_duration(end_lon, end_lat, base_lon, base_lat)
                    if ret is None: ret = estimate_travel_time_fallback(haversine(end_lon, end_lat, base_lon, base_lat))
                    
                    if total + ret <= DAILY_LIMIT_MINS:
                        tasks[idx]['Assigned'] = True
                        state['time_used'] = total
                        state['curr_lon'] = end_lon
                        state['curr_lat'] = end_lat
                        state['curr_county'] = '桃園市' # PZ-B is Taoyuan
                        state['tasks_done'] += 1
                        
                        full_schedule.append({
                            'Day': day,
                            'Staff_ID': s_id,
                            'Zone': t['Zone'],
                            'Seq': state['tasks_done'],
                            'Task_IDs': t['ID_List'],
                            'Address': t['Address'],
                            'Lat': t['Lat'],
                            'Lon': t['Lon'],
                            'Work_Mins': work_time,
                            'Travel_Mins': round(leg, 1),
                            'Total_Time': round(state['time_used'], 1),
                            'Job_Type': 'Primary (Priority)'
                        })
                        p_change = True
                        break

        
        change_made = True
        while change_made:
            change_made = False
            
            for staff in STAFF_CONFIG:
                s_id = staff['id']
                state = staff_state[s_id]
                
                # Check if full
                if state['time_used'] >= DAILY_LIMIT_MINS:
                    continue
                
                # Find eligible zones
                eligible_zones = staff['zones']
                
                # 1. Collect ALL eligible tasks
                candidates = []
                for i, t in enumerate(tasks):
                    if not t['Assigned'] and t['Zone'] in eligible_zones:
                        # Use Haversine for initial sorting to reduce API calls
                        dist = haversine(state['curr_lon'], state['curr_lat'], t['Lon'], t['Lat'])
                        candidates.append((dist, i))
                
                # 2. Sort by distance (Nearest first)
                candidates.sort(key=lambda x: x[0])
                
                # 3. Find the first one that fits (Greedy with OSRM verification)
                found_task = False
                
                # Optimization: Only check top 10 nearest to avoid excessive API calls per step
                # If cannot find in top 10, maybe skip? Or check longer?
                # Let's check top 5.
                for _, idx in candidates[:5]: 
                    t = tasks[idx]
                    
                    # --- Determine Routing Logic ---
                    target_county = get_county(t['Address'])
                    base_key = staff['team']
                    base_lon = base_info[base_key]['lon']
                    base_lat = base_info[base_key]['lat']
                    
                    duration_mins = 0
                    travel_path_desc = "Direct"
                    
                    # Logic: If Cross County -> Go via Base
                    # Condition: If current county is NOT the same as target county
                    # Special check: If we are AT base, current county IS base county.
                    # If target is different, we just go direct (Base -> Target).
                    # We only need to "Return to Base" if we are at a Text A (County A) and go to Task B (County B)
                    # and County A != County B.
                    
                    # BUT wait, simply checking county equality might be too strict if they are adjacent?
                    # User said: "If cross county... return to warehouse first".
                    # We will follow strictly.
                    
                    is_cross_county = (state['curr_county'] != target_county) and (state['curr_county'] != 'Unknown') and (target_county != 'Unknown')
                    
                    # Exception: If I am AT Base (start of day), I am in Base County.
                    # If I go to another county, I go direct. The rule "Return to Base" implies I am NOT at Base.
                    # So if tasks_done == 0, is_cross_county is False (effectively).
                    if state['tasks_done'] == 0:
                        is_cross_county = False
                        
                    # Also if I am ALREADY at base (maybe returned previously?)
                    # Distance check to base?
                    dist_from_base = haversine(state['curr_lon'], state['curr_lat'], base_lon, base_lat)
                    if dist_from_base < 0.2: # 200m
                         is_cross_county = False
                    
                    if is_cross_county:
                        # Detour via base
                        # 1. Current -> Base
                        leg1 = get_osrm_duration(state['curr_lon'], state['curr_lat'], base_lon, base_lat)
                        if leg1 is None: leg1 = estimate_travel_time_fallback(dist_from_base)
                        
                        # 2. Base -> Target
                        dist_base_target = haversine(base_lon, base_lat, t['Lon'], t['Lat'])
                        leg2 = get_osrm_duration(base_lon, base_lat, t['Lon'], t['Lat'])
                        if leg2 is None: leg2 = estimate_travel_time_fallback(dist_base_target)
                        
                        duration_mins = leg1 + leg2
                        travel_path_desc = "Via Base"
                    else:
                        # Direct
                        dist_direct = haversine(state['curr_lon'], state['curr_lat'], t['Lon'], t['Lat'])
                        leg = get_osrm_duration(state['curr_lon'], state['curr_lat'], t['Lon'], t['Lat'])
                        if leg is None: leg = estimate_travel_time_fallback(dist_direct)
                        duration_mins = leg
                        
                    # --- End Routing Logic ---
                    
                    work_time = t['Work_Mins']
                    total_time_at_finish = state['time_used'] + duration_mins + work_time
                    
                    if total_time_at_finish <= DAILY_LIMIT_MINS:
                        # Check Return to Base at END of day
                        # We must ensure there is enough time to return to base from the new task
                        # Is end-of-day return considered "Cross County"?
                        # Usually yes, but the valid time check usually just adds the direct trip home.
                        # OR if the return trip crosses county, do we need to double return? No, target IS base.
                        
                        dist_return = haversine(t['Lon'], t['Lat'], base_lon, base_lat)
                        return_home_time = get_osrm_duration(t['Lon'], t['Lat'], base_lon, base_lat)
                        if return_home_time is None: return_home_time = estimate_travel_time_fallback(dist_return)
                        
                        if total_time_at_finish + return_home_time <= DAILY_LIMIT_MINS:
                            # Valid Assignment
                            tasks[idx]['Assigned'] = True
                            state['time_used'] = total_time_at_finish
                            state['curr_lon'] = t['Lon']
                            state['curr_lat'] = t['Lat']
                            state['curr_county'] = target_county
                            state['tasks_done'] += 1
                            
                            full_schedule.append({
                                'Day': day,
                                'Staff_ID': s_id,
                                'Zone': t['Zone'],
                                'Seq': state['tasks_done'],
                                'Task_IDs': t['ID_List'],
                                'Address': t['Address'],
                                'Lat': t['Lat'],
                                'Lon': t['Lon'],
                                'Work_Mins': work_time,
                                'Travel_Mins': round(duration_mins, 1),
                                'Total_Time': round(state['time_used'], 1),
                                'Job_Type': 'Primary' if travel_path_desc == "Direct" else 'Primary (Via Base)'
                            })
                            
                            found_task = True
                            change_made = True
                            # time.sleep(0.05) # Rate limit protection
                            break
                        
                if not found_task:
                    pass

        # Support Phase (Cross-Zone) was intentionally removed in Phase 2 to prioritize locality.
        # Staff now only work within their assigned zones.
        pass
                        
    # End of week
    unassigned = [t for t in tasks if not t['Assigned']]
    print(f"Scheduling Complete. Unassigned Tasks: {len(unassigned)}")
    if unassigned:
        print("--- Sample Unassigned Tasks ---")
        for t in unassigned[:10]:
            print(f"Zone: {t['Zone']}, Work: {t['Work_Mins']}, Addr: {t['Address']}")
        pd.DataFrame(unassigned).to_csv('unassigned_tasks.csv', index=False, encoding='utf-8-sig')
    
    # Save with Color Coding using openpyxl
    print(f"Saving to {OUTPUT_FILE} with color coding...")
    df_res = pd.DataFrame(full_schedule)
    
    # Sort for Clarity
    df_res = df_res.sort_values(by=['Staff_ID', 'Day', 'Seq'])
    
    # Capture Staff IDs for logic before renaming
    staff_ids = sorted(list(set(df_res['Staff_ID'])))
    
    # Rename Columns for User Friendliness (Chinese)
    col_map = {
        'Day': '日程(Day)', 
        'Staff_ID': '員工代號', 
        'Zone': '工作區域', 
        'Seq': '順序', 
        'Task_IDs': '任務ID', 
        'Address': '地址', 
        'Lat': '緯度',
        'Lon': '經度',
        'Work_Mins': '維護時間(分)', 
        'Travel_Mins': '預估車程(分)', 
        'Total_Time': '累計工時(分)', 
        'Job_Type': '任務屬性'
    }
    df_res = df_res.rename(columns=col_map)
    
    df_res.to_excel(OUTPUT_FILE, index=False)
    
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        
        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active
        
        # 14 Distinct Pastel Colors (Hex)
        colors = [
            'FF9999', '99FF99', '9999FF', 'FFFF99', 'FF99FF', '99FFFF', 'FFCC99', 
            'CCCCFF', 'CCFFCC', 'FFCCCC', 'E0E0E0', 'FFD700', 'ADFF2F', '00BFFF'
        ]
        
        # staff_ids computed above
        staff_color_map = {s_id: colors[i % len(colors)] for i, s_id in enumerate(staff_ids)}
        support_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid') # Orange
        
        # Identify Columns by Header (Row 1)
        header = {cell.value: i+1 for i, cell in enumerate(ws[1])} # Name -> 1-based index
        
        staff_col_idx = header.get('員工代號')
        job_col_idx = header.get('任務屬性')
        
        if not staff_col_idx or not job_col_idx:
            print(f"Warning: Columns not found. Header: {header}")
        else:
            # Iterate through all data rows
            for row in range(2, ws.max_row + 1):
                # 1. Staff Color
                staff_cell = ws.cell(row=row, column=staff_col_idx)
                s_id = staff_cell.value
                if s_id in staff_color_map:
                    fill = PatternFill(start_color=staff_color_map[s_id], end_color=staff_color_map[s_id], fill_type='solid')
                    staff_cell.fill = fill
                    
                # 2. Support Color
                job_cell = ws.cell(row=row, column=job_col_idx)
                if 'Support' in str(job_cell.value):
                    job_cell.fill = support_fill
                    
        wb.save(OUTPUT_FILE)
        print("Done.")
    except Exception as e:
        print(f"Coloring failed: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    generate_schedule()
