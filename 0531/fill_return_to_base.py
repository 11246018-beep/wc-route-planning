import openpyxl
from openpyxl.styles import PatternFill
import requests
import json
import os
import time

# Configuration
INPUT_FILE = 'Weekly_Schedule_Updated.xlsx'
OUTPUT_FILE = 'Weekly_Schedule_Updated.xlsx' # Overwrite or same name
CACHE_FILE = 'osrm_cache.json'
OSRM_BASE_URL = 'http://router.project-osrm.org/route/v1/driving'

# Base Coordinates
# PZ: Pingzhen, WG: Wugu
BASE_COORDS = {'PZ': (24.90679, 121.22683), 'WG': (25.07055, 121.44141)}
STAFF_TEAM_MAP = {
    'S01': 'PZ', 'S02': 'PZ', 'S03': 'PZ', 'S04': 'PZ', 'S05': 'PZ', 'S06': 'PZ', 'S07': 'PZ', 
    'S08': 'PZ', 'S09': 'PZ', 'S10': 'PZ', 'S11': 'PZ', 'S12': 'PZ', 'S13': 'WG', 'S14': 'WG'
}

# Cache Handling
osrm_cache = {}

def load_cache():
    global osrm_cache
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                osrm_cache = json.load(f)
            # print(f"Loaded {len(osrm_cache)} entries from cache.")
        except Exception as e:
            print(f"Error loading cache: {e}", flush=True)

def save_cache():
    try:
        with open(CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(osrm_cache, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"Error saving cache: {e}", flush=True)

def get_osrm_duration(start_lat, start_lon, end_lat, end_lon):
    """
    Returns duration in minutes between two points using OSRM API.
    Uses caching.
    """
    r_start_lat, r_start_lon = round(start_lat, 4), round(start_lon, 4)
    r_end_lat, r_end_lon = round(end_lat, 4), round(end_lon, 4)

    key = f"{r_start_lat},{r_start_lon}|{r_end_lat},{r_end_lon}"
    
    if key in osrm_cache:
        return osrm_cache[key]
    
    try:
        url = f"{OSRM_BASE_URL}/{r_start_lon},{r_start_lat};{r_end_lon},{r_end_lat}?overview=false"
        response = requests.get(url, timeout=5)
        
        if response.status_code == 200:
            data = response.json()
            if 'routes' in data and len(data['routes']) > 0:
                duration_sec = data['routes'][0]['duration']
                duration_mins = duration_sec / 60.0
                osrm_cache[key] = duration_mins
                return duration_mins
            else:
                return 0
        else:
            return 0
    except Exception as e:
        print(f"Request Error for {key}: {e}", flush=True)
        return 0

def fill_return_to_base():
    load_cache()
    
    print(f"Loading {INPUT_FILE}...", flush=True)
    # Load workbook preserving styles
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active
    
    # 1. Identify Columns
    headers = {cell.value: i+1 for i, cell in enumerate(ws[1])}
    
    col_staff = headers.get('員工代號')
    col_lat = headers.get('緯度')
    col_lon = headers.get('經度')
    col_addr = headers.get('地址')
    col_time = headers.get('抵達所需時間')
    
    if not all([col_staff, col_lat, col_lon, col_time]):
        print("Required columns missing.", flush=True)
        return

    print("Processing Yellow Rows...", flush=True)
    count = 0
    updated = 0
    
    # Iterate rows
    for row in range(2, ws.max_row + 1):
        cell_time = ws.cell(row=row, column=col_time)
        fill = cell_time.fill
        
        # Check Yellow
        if fill and fill.patternType == 'solid':
            fg = fill.start_color.rgb
            if fg and 'FFFF00' in str(fg): 
                # Found Target Row
                # 1. Get Staff
                staff_id = ws.cell(row=row, column=col_staff).value
                
                # 2. Get Previous Row (Origin)
                # Need to find the previous valid row for this staff/day?
                # Or just literally the row above?
                # User inserted "in between", so row-1 should be the "Previous Job".
                # But careful if row-1 is header or empty.
                prev_row = row - 1
                if prev_row < 2: continue
                
                prev_lat = ws.cell(row=prev_row, column=col_lat).value
                prev_lon = ws.cell(row=prev_row, column=col_lon).value
                
                if not prev_lat or not prev_lon:
                    print(f"Row {row}: Previous row has no coords.", flush=True)
                    continue
                    
                # 3. Get Base Coords
                team = STAFF_TEAM_MAP.get(staff_id, 'PZ')
                base_lat, base_lon = BASE_COORDS[team]
                
                # 4. Calculate
                duration = get_osrm_duration(prev_lat, prev_lon, base_lat, base_lon)
                
                # 5. Update
                cell_time.value = round(duration, 1)
                
                # Optional: Set Address to Base Name if empty?
                # cell_addr = ws.cell(row=row, column=col_addr)
                # if not cell_addr.value:
                #     cell_addr.value = f"返回{team}倉庫"
                
                print(f"Row {row}: Updated time to {duration:.1f} min (Staff {staff_id})", flush=True)
                updated += 1
                
                if updated % 5 == 0:
                    save_cache()

    save_cache()
    wb.save(OUTPUT_FILE)
    print(f"Done. Updated {updated} rows in {OUTPUT_FILE}", flush=True)

if __name__ == "__main__":
    fill_return_to_base()
