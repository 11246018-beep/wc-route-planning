import openpyxl
from openpyxl.styles import PatternFill

INPUT_FILE = 'Weekly_Schedule_Updated.xlsx'

def revert_yellow_cells():
    print(f"Loading {INPUT_FILE}...", flush=True)
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active
    
    # 1. Identify Columns
    headers = {cell.value: i+1 for i, cell in enumerate(ws[1])}
    col_time = headers.get('抵達所需時間')
    
    if not col_time:
        print("Column '抵達所需時間' not found.", flush=True)
        return

    print("Reverting Yellow Rows (Clearing values)...", flush=True)
    reverted_count = 0
    
    # Iterate rows
    for row in range(2, ws.max_row + 1):
        cell_time = ws.cell(row=row, column=col_time)
        fill = cell_time.fill
        
        # Check Yellow
        if fill and fill.patternType == 'solid':
            fg = fill.start_color.rgb
            if fg and 'FFFF00' in str(fg): 
                # Found Target Row - Clear Value
                if cell_time.value is not None:
                    cell_time.value = None
                    reverted_count += 1
                    # print(f"Row {row}: Cleared value.", flush=True)

    wb.save(INPUT_FILE)
    print(f"Done. Reverted (Cleared) {reverted_count} rows in {INPUT_FILE}", flush=True)

if __name__ == "__main__":
    revert_yellow_cells()
