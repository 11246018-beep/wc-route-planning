import openpyxl
from openpyxl.styles import PatternFill

INPUT_FILE = 'Weekly_Schedule_Updated.xlsx'

def inspect_colors():
    print(f"Loading {INPUT_FILE}...")
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
    ws = wb.active
    
    # Find headers
    headers = {cell.value: i+1 for i, cell in enumerate(ws[1])}
    target_col_idx = headers.get('抵達所需時間')
    
    if not target_col_idx:
        print("Column '抵達所需時間' not found.")
        return

    print(f"Target Column Index: {target_col_idx}")
    
    found_yellow = 0
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=target_col_idx)
        fill = cell.fill
        
        # Check for non-none fill
        if fill and fill.patternType == 'solid':
            fg = fill.start_color.rgb
            # print(f"Row {row} Color: {fg}")
            
            # Common Yellows: FFFF0000 (Red?), FFFFFF00 (Yellow), 00FFFF00
            # OpenPyXL colors can be a bit tricky (ARGB). Yellow is often FFFFFF00.
            if fg and 'FFFF00' in str(fg): 
                found_yellow += 1
                print(f"Row {row}: Found Yellow! Value: {cell.value}")
                
    print(f"Total Yellow Rows Found: {found_yellow}")

if __name__ == "__main__":
    inspect_colors()
