import pandas as pd
import numpy as np
from math import radians, cos, sin, asin, sqrt

# Configuration
INPUT_FILE = 'maintenance_data_v2.xlsx'
OUTPUT_FILE = 'Weekly_Schedule.xlsx'

# Staff Configuration
STAFF_CONFIG = [
    {'id': 'S01', 'team': 'PZ', 'zones': ['PZ-A']},
    {'id': 'S02', 'team': 'PZ', 'zones': ['PZ-A']},
    {'id': 'S03', 'team': 'PZ', 'zones': ['PZ-A']},
    {'id': 'S04', 'team': 'PZ', 'zones': ['PZ-A']}, # 4 Staff (8.4k mins) -> ~2100 mins each
    {'id': 'S05', 'team': 'PZ', 'zones': ['PZ-B']},
    {'id': 'S06', 'team': 'PZ', 'zones': ['PZ-B']},
    {'id': 'S07', 'team': 'PZ', 'zones': ['PZ-B']}, # 3 Staff (7.2k mins) -> ~2400 mins each
    {'id': 'S08', 'team': 'PZ', 'zones': ['PZ-D']},
    {'id': 'S09', 'team': 'PZ', 'zones': ['PZ-D']},
    {'id': 'S10', 'team': 'PZ', 'zones': ['PZ-D']}, # 3 Staff (7.2k mins) -> ~2400 mins each
    {'id': 'S11', 'team': 'PZ', 'zones': ['PZ-C']},
    {'id': 'S12', 'team': 'PZ', 'zones': ['PZ-C']}, # 2 Staff (5.5k mins) -> ~2700 mins each
    {'id': 'S13', 'team': 'WG', 'zones': ['WG-A', 'WG-B', 'WG-C', 'WG-D']},
    {'id': 'S14', 'team': 'WG', 'zones': ['WG-A', 'WG-B', 'WG-C', 'WG-D']} # 2 Staff (3k mins) -> Very light load, maybe support PZ-D?
]

DAILY_LIMIT_MINS = 540 # Strict limit requested by user
AVG_SPEED_KMPH = 25  # Conservative city speed
MIN_TRAVEL_TIME = 10 # (Unused variable now)

def haversine(lon1, lat1, lon2, lat2):
    """
    Calculate the great circle distance in kilometers between two points 
    on the earth (specified in decimal degrees)
    """
    if pd.isna(lon1) or pd.isna(lat1) or pd.isna(lon2) or pd.isna(lat2):
        return 0
    # convert decimal degrees to radians 
    lon1, lat1, lon2, lat2 = map(radians, [lon1, lat1, lon2, lat2])
    # haversine formula 
    dlon = lon2 - lon1 
    dlat = lat2 - lat1 
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * asin(sqrt(a)) 
    r = 6371 # Radius of earth in kilometers
    return c * r

def estimate_travel_time(km):
    if km < 0.1: return 2 # Elevator/Walking time for same building
    
    # Speed Model
    # Short trips (City): 25 km/h
    # Long trips (Highway): 60 km/h (conservative highway avg including exit/entry)
    speed = AVG_SPEED_KMPH
    if km > 10:
        speed = 60
        
    hours = km / speed
    mins = hours * 60
    return max(mins, 3) # Reducing buffer to 3 mins

def get_zone(sheet, addr):
    addr = str(addr)
    # Pingzhen Logic
    if '平鎮' in sheet:
        # Priority 1: Explicit Zone Labels
        a_sub = ['大園', '蘆竹', '觀音', '桃園區', '龜山']
        b_sub = ['中壢', '平鎮', '八德', '龍潭', '楊梅', '新屋']
        c_sub = ['新竹', '竹北', '湖口', '新豐', '寶山', '竹東', '芎林', '新埔', '關西', '峨眉']
        d_sub = ['苗栗', '新北', '臺北', '台北', '頭份', '竹南', '公館', '造橋']

        # Check in order of specificity. 
        # Check B first? Or check specific '桃園區' to avoid '桃園市'.
        
        if any(k in addr for k in c_sub): return 'PZ-C'
        if any(k in addr for k in d_sub): return 'PZ-D'
        if any(k in addr for k in b_sub): return 'PZ-B'
        if any(k in addr for k in a_sub): return 'PZ-A'
        
        # If '桃園' in addr but NO district matched above? 
        # E.g. '桃園市xx路'. Likely A (Taoyuan Dist) or B. 
        # But '桃園市' is weak signal.
        if '桃園' in addr and '區' not in addr: return 'PZ-A' # Fallback
        
        return 'PZ-B' # Ultimate Default (Central Area)
    # Wugu Logic
    elif '五股' in sheet:
        if any(k in addr for k in ['北投', '淡水', '八里', '三芝', '金山', '石門']): return 'WG-B'
        if any(k in addr for k in ['基隆', '汐止', '瑞芳', '七堵', '安樂', '中山', '仁愛', '信義', '中正']): return 'WG-C'
        if any(k in addr for k in ['桃園', '龜山']): return 'WG-D'
        return 'WG-A' # Default
    return 'Unknown'

def load_tasks():
    print("Loading tasks...")
    all_tasks = []
    
    # Process each sheet
    # We rely on ID being present (Column 0)
    for sheet in ['平鎮', '平鎮週清2', '五股', '五股週清2']:
        try:
            df = pd.read_excel(INPUT_FILE, sheet_name=sheet)
            
            # Identify columns
            id_col = 'ID'
            addr_col = next((c for c in df.columns if '地址' in str(c) or '地點' in str(c)), None)
            lat_col = next((c for c in df.columns if '緯度' in str(c)), None)
            lon_col = next((c for c in df.columns if '經度' in str(c)), None)
            floor_col = next((c for c in df.columns if '樓層' in str(c)), None) # Usually explicit in Master not here?
            
            # Floor info is in Master mainly. Let's try to map from Master if needed?
            # Or assume floor data is missing here and re-merge?
            # User said "put ID into these sheets".
            # The Pingzhen sheet looks like: ['ID', '地址', '加總 - 維護時間'...]
            # '加總 - 維護時間' likely IS the workload? 
            # Or we can use the "floor * 8" rule if floor is missing.
            # Let's check if '樓層' exists or '維護時間' exists.
            
            work_col = next((c for c in df.columns if '維護時間' in str(c)), None)
            
            # Drop rows where ID is NaN (avoids Summary rows)
            df = df.dropna(subset=[id_col])
            
            # If lat/lon missing, we can arguably skip or warn.
            # But definitely drop summary rows first.
            if lat_col and lon_col:
                df = df.dropna(subset=[lat_col, lon_col])
            
            for _, row in df.iterrows():
                # Handle comma separated IDs
                ids = str(row[id_col]).split(',')
                # If multiple IDs share one location, strictly they are multiple tasks.
                # BUT for routing they are one stop.
                # Splitting them allows assigning to different people/days BUT same location.
                # Let's split them.
                
                # Workload per ID?
                # If '維護時間' is aggregated for the address, we should split it?
                # Or assigns the whole block?
                # Let's assume the row represents one "Stop" which may contain multiple "Tasks".
                # For scheduling, it's easier to schedule the Stop.
                
                total_work = row[work_col] if work_col else 30 # Default 30 if missing
                
                # Determine Zone
                if '工作區' in df.columns and pd.notna(row['工作區']):
                    # Clean up 'PZ-A 北桃園' to 'PZ-A'
                    z_raw = str(row['工作區'])
                    if 'PZ-A' in z_raw: zone = 'PZ-A'
                    elif 'PZ-B' in z_raw: zone = 'PZ-B'
                    elif 'PZ-C' in z_raw: zone = 'PZ-C'
                    elif 'PZ-D' in z_raw: zone = 'PZ-D'
                    else: zone = get_zone(sheet, row[addr_col])
                else:
                    zone = get_zone(sheet, row[addr_col])
                
                all_tasks.append({
                    'ID_List': row[id_col],
                    'Address': row[addr_col],
                    'Lat': row[lat_col],
                    'Lon': row[lon_col],
                    'Work_Mins': total_work,
                    'Zone': zone,
                    'Sheet': sheet,
                    'Assigned': False
                })
                
        except Exception as e:
            print(f"Error loading {sheet}: {e}")
            
    print(f"Loaded {len(all_tasks)} stops.")
    total_work = sum(t['Work_Mins'] for t in all_tasks)
    print(f"Total Workload to Schedule: {total_work} mins")
    return all_tasks

def generate_schedule():
    tasks = load_tasks()
    schedule = []
    
    # Bases
    # Bases
    # Load from '倉位地點' sheet
    base_coords = {'PZ': (121.22683, 24.90679), 'WG': (121.44141, 25.07055)} # Default fallback
    try:
        df_wh = pd.read_excel(INPUT_FILE, sheet_name='倉位地點')
        # Assuming format: Unnamed: 0 (Name), 經度, 緯度
        # Identify columns
        name_col = df_wh.columns[0]
        lon_col = next((c for c in df_wh.columns if '經度' in str(c)), None)
        lat_col = next((c for c in df_wh.columns if '緯度' in str(c)), None)
        
        if lon_col and lat_col:
            for _, row in df_wh.iterrows():
                name = str(row[name_col])
                if '平鎮' in name:
                    base_coords['PZ'] = (row[lon_col], row[lat_col])
                elif '五股' in name:
                    base_coords['WG'] = (row[lon_col], row[lat_col])
            print(f"Loaded Base Coords: {base_coords}")
    except Exception as e:
        print(f"Warning: Could not load warehouse sheet, using defaults. {e}")
    
    # 1. Distribute tasks to Staff based on Zone
    # Create Staff Queues
    staff_queues = {s['id']: [] for s in STAFF_CONFIG}
    
    # Sort tasks into Zone Pools
    zone_pools = {} # {'PZ-A': [tasks], ...}
    for t in tasks:
        z = t['Zone']
        if z not in zone_pools: zone_pools[z] = []
        zone_pools[z].append(t)
        
    # Assign Pools to Staff
    # We round-robin assign tasks from Zone Pools to eligible Staff
    
    for staff in STAFF_CONFIG:
        s_id = staff['id']
        s_zones = staff['zones']
        
        # Collect all eligible tasks
        # But wait, tasks should be shared among staff in same zone.
        pass
        
    # Better approach:
    # Iterate Days 1-6.
    # For each Day, iterate Staff.
    # Staff picks nearest task from their eligible Zone Pool.
    
    full_schedule = []
    
    for day in range(1, 7):
        print(f"--- Scheduling Day {day} ---")
        
        # Reset Staff State for new Day
        staff_state = {}
        for s in STAFF_CONFIG:
            team = s['team']
            start_lon, start_lat = base_coords[team]
            staff_state[s['id']] = {
                'curr_lon': start_lon,
                'curr_lat': start_lat,
                'time_used': 0,
                'tasks_done': 0
            }
            
        # Optimization: We loop unassigned tasks until no one can take more
        # But iterating all 2000 tasks every step is slow.
        # We should filter tasks by 'Assigned' = False.
        
        # To avoid infinite loop if tasks don't fit
        change_made = True
        while change_made:
            change_made = False
            
            for staff in STAFF_CONFIG:
                s_id = staff['id']
                state = staff_state[s_id]
                
                # Check if full
                if state['time_used'] >= DAILY_LIMIT_MINS:
                    continue
                
                # Find eligible zones
                eligible_zones = staff['zones']
                
                # Improved Logic: 
                # 1. Collect ALL eligible tasks for this staff
                candidates = []
                for i, t in enumerate(tasks):
                    if not t['Assigned'] and t['Zone'] in eligible_zones:
                        dist = haversine(state['curr_lon'], state['curr_lat'], t['Lon'], t['Lat'])
                        candidates.append((dist, i))
                
                # 2. Sort by distance (Nearest first)
                candidates.sort(key=lambda x: x[0])
                
                # 3. Find the first one that fits
                found_task = False
                for dist, idx in candidates:
                    t = tasks[idx]
                    travel_time = estimate_travel_time(dist)
                    work_time = t['Work_Mins']
                    total_time = travel_time + work_time
                    
                    if state['time_used'] + total_time <= DAILY_LIMIT_MINS:
                        # Check return trip logic
                        base_lon, base_lat = base_coords[staff['team']]
                        dist_to_base = haversine(t['Lon'], t['Lat'], base_lon, base_lat)
                        return_time = estimate_travel_time(dist_to_base)
                        
                        if state['time_used'] + total_time + return_time <= DAILY_LIMIT_MINS:
                            # Assign
                            tasks[idx]['Assigned'] = True
                            state['time_used'] += total_time
                            state['curr_lon'] = t['Lon']
                            state['curr_lat'] = t['Lat']
                            state['tasks_done'] += 1
                            
                            full_schedule.append({
                                'Day': day,
                                'Staff_ID': s_id,
                                'Zone': t['Zone'],
                                'Seq': state['tasks_done'],
                                'Task_IDs': t['ID_List'],
                                'Address': t['Address'],
                                'Lat': t['Lat'],
                                'Lon': t['Lon'],
                                'Work_Mins': work_time,
                                'Travel_Mins': round(travel_time, 1),
                                'Total_Time': round(state['time_used'], 1),
                                'Job_Type': 'Primary'
                            })
                            
                            found_task = True
                            change_made = True
                            break # Move to next staff (Round Robin-ish inside greedy loop) or stay with this staff?
                            # If we break here, we go to next staff. This balances load.
                        
                if not found_task:
                    # No more tasks fit this staff for the PRIMARY zones
                    pass

        # --- Support Phase (Cross-Zone) ---
        # Allow staff to pick up ANY unassigned task if they have time
        print(f"  > Day {day} Support Phase...")
        
        change_made = True
        while change_made:
            change_made = False
            for staff in STAFF_CONFIG:
                s_id = staff['id']
                state = staff_state[s_id]
                
                if state['time_used'] >= DAILY_LIMIT_MINS: continue
                
                # Find nearest GLOBALLY unassigned task
                candidates = []
                for i, t in enumerate(tasks):
                    if not t['Assigned']:
                        dist = haversine(state['curr_lon'], state['curr_lat'], t['Lon'], t['Lat'])
                        candidates.append((dist, i))
                        
                # Prioritize PZ-D tasks (Miaoli) specifically for Support
                # to ensure they are done when time block is large enough.
                candidates.sort(key=lambda x: (
                    0 if tasks[x[1]]['Zone'] == 'PZ-D' else 1, 
                    x[0]
                ))
                
                for dist, idx in candidates:
                    t = tasks[idx]
                    travel_time = estimate_travel_time(dist)
                    work_time = t['Work_Mins']
                    total_time = travel_time + work_time
                    
                    if state['time_used'] + total_time <= DAILY_LIMIT_MINS:
                        # Check return trip logic
                        base_lon, base_lat = base_coords[staff['team']]
                        dist_to_base = haversine(t['Lon'], t['Lat'], base_lon, base_lat)
                        return_time = estimate_travel_time(dist_to_base)

                        if state['time_used'] + total_time + return_time <= DAILY_LIMIT_MINS:
                            # Assign as Support
                            tasks[idx]['Assigned'] = True
                            state['time_used'] += total_time
                            state['curr_lon'] = t['Lon']
                            state['curr_lat'] = t['Lat']
                            state['tasks_done'] += 1
                            
                            full_schedule.append({
                                'Day': day,
                                'Staff_ID': s_id,
                                'Zone': t['Zone'],
                                'Seq': state['tasks_done'],
                                'Task_IDs': t['ID_List'],
                                'Address': t['Address'],
                                'Lat': t['Lat'],
                                'Lon': t['Lon'],
                                'Work_Mins': work_time,
                                'Travel_Mins': round(travel_time, 1),
                                'Total_Time': round(state['time_used'], 1),
                                'Job_Type': 'Support' # Mark as Support
                            })
                            change_made = True
                            break # Move to next staff for load balancing
                        
    # End of week
    unassigned = [t for t in tasks if not t['Assigned']]
    print(f"Scheduling Complete. Unassigned Tasks: {len(unassigned)}")
    if unassigned:
        print("--- Sample Unassigned Tasks ---")
        for t in unassigned[:10]:
            print(f"Zone: {t['Zone']}, Work: {t['Work_Mins']}, Addr: {t['Address']}")
        
        # Expert Unassigned to CSV for analysis
        pd.DataFrame(unassigned).to_csv('unassigned_tasks.csv', index=False, encoding='utf-8-sig')
    
    # Save with Color Coding using openpyxl
    print(f"Saving to {OUTPUT_FILE} with color coding...")
    df_res = pd.DataFrame(full_schedule)
    
    # Sort for Clarity (User Request: List routes by Staff)
    df_res = df_res.sort_values(by=['Staff_ID', 'Day', 'Seq'])
    
    # Capture Staff IDs for logic before renaming
    staff_ids = sorted(list(set(df_res['Staff_ID'])))
    
    # Rename Columns for User Friendliness (Chinese)
    col_map = {
        'Day': '日程(Day)', 
        'Staff_ID': '員工代號', 
        'Zone': '工作區域', 
        'Seq': '順序', 
        'Task_IDs': '任務ID', 
        'Address': '地址', 
        'Lat': '緯度',
        'Lon': '經度',
        'Work_Mins': '維護時間(分)', 
        'Travel_Mins': '預估車程(分)', 
        'Total_Time': '累計工時(分)', 
        'Job_Type': '任務屬性'
    }
    df_res = df_res.rename(columns=col_map)
    
    df_res.to_excel(OUTPUT_FILE, index=False)
    
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill
        
        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active
        
        # 14 Distinct Pastel Colors (Hex)
        colors = [
            'FF9999', '99FF99', '9999FF', 'FFFF99', 'FF99FF', '99FFFF', 'FFCC99', 
            'CCCCFF', 'CCFFCC', 'FFCCCC', 'E0E0E0', 'FFD700', 'ADFF2F', '00BFFF'
        ]
        
        # staff_ids computed above
        staff_color_map = {s_id: colors[i % len(colors)] for i, s_id in enumerate(staff_ids)}
        support_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid') # Orange
        
        # Identify Columns by Header (Row 1)
        header = {cell.value: i+1 for i, cell in enumerate(ws[1])} # Name -> 1-based index
        
        staff_col_idx = header.get('員工代號')
        job_col_idx = header.get('任務屬性')
        
        if not staff_col_idx or not job_col_idx:
            print(f"Warning: Columns not found. Header: {header}")
        else:
            # Iterate through all data rows
            for row in range(2, ws.max_row + 1):
                # 1. Staff Color
                staff_cell = ws.cell(row=row, column=staff_col_idx)
                s_id = staff_cell.value
                if s_id in staff_color_map:
                    fill = PatternFill(start_color=staff_color_map[s_id], end_color=staff_color_map[s_id], fill_type='solid')
                    staff_cell.fill = fill
                    
                # 2. Support Color
                job_cell = ws.cell(row=row, column=job_col_idx)
                if job_cell.value == 'Support':
                    job_cell.fill = support_fill
                    
        wb.save(OUTPUT_FILE)
        print("Done.")
    except Exception as e:
        print(f"Coloring failed: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    generate_schedule()
